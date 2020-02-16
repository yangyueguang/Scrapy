# coding! utf-8
# 这是管道文件
import os
import time
import json
import redis
import pymongo
import pymysql
import requests
import datetime
import openpyxl
from project import conf
from openpyxl.styles import PatternFill, colors


class YYPipeline(object):

    def __init__(self):
        self.se = requests.session()
        self.se.headers = {
            'Accept': '*/*',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'en-US,en;q=0.9,zh-CN;q=0.8,zh;q=0.7',
            'Connection': 'keep-alive',
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36',
        }
        self.excel_file = conf.excel_file
        sheet_name = '记录'
        if not os.path.isfile(self.excel_file):
            self.workbook = openpyxl.Workbook()
            self.workbook.remove(self.workbook['Sheet'])
        else:
            self.workbook = openpyxl.load_workbook(self.excel_file)
        if sheet_name in self.workbook.sheetnames:
            self.sheet = self.workbook[sheet_name]
        else:
            self.sheet = self.workbook.create_sheet(sheet_name)
        self.sheet.cell(1, 1, '项目名称')
        self.sheet.cell(1, 2, '采购单位')
        self.sheet.cell(1, 3, '发布时间')
        self.sheet.cell(1, 4, '原文地址')
        self.sheet.cell(1, 5, '数据来源')
        self.workbook.save(self.excel_file)

    def process_item(self, item, spider):
        row = self.sheet.max_row + 1
        if self.verify_useful(item):
            self.sheet.cell(row, 1, item['name'])
            self.sheet.cell(row, 2, item['unit'])
            self.sheet.cell(row, 3, item['time'])
            self.sheet.cell(row, 4, item['address'])
            self.sheet.cell(row, 5, item['sources'])
            return item
        else:
            return None

    def close_spider(self, spider):
        self.workbook.save(self.excel_file)

    # 验证文章是否是符合需求的
    def verify_useful(self, item):
        time.sleep(2)
        # return True
        # TODO
        date = datetime.datetime.strptime(item['time'], '%Y-%m-%d %H:%M')
        n = datetime.datetime.now().date()
        now_zero = datetime.datetime.now().replace(year=n.year, month=n.month, day=n.day, hour=0, minute=0, second=0)
        day = (now_zero - date).days
        # 包含昨天和今天的
        if day > 1 or day < -1:
            return False
        source = item['sources']
        if source == '剑鱼网':
            return True
        elif source == '金采网':
            source_id = item['address'].split('=')[-1]
            detail_url = 'http://www.cfcpn.com/jcw/noticeinfo/noticeInfo/dataNoticeList'
            res = self.se.post(detail_url, data={'id': source_id, 'isDetail': '1'})
            res.encoding = 'utf-8'
        elif source == '中国东方航空采购招标网':
            source_id = item['address'].split('=')[-1]
            detail_url = 'https://caigou.ceair.com/portal/portal/findSysInfo?siteFlag=false'
            res = self.se.post(detail_url, data={'id': source_id})
            res.encoding = 'utf-8'
        elif source == '中国联通':
            self.se.get('http://www.chinaunicombidding.cn/jsp/cnceb/web/index_parent.jsp')
            res = self.se.get(item['address'])
            res.encoding = 'gb2312'
        else:
            res = self.se.get(item['address'])
            res.encoding = 'utf-8'
        print(res.url)
        print(res.status_code)
        if res.status_code != 200:
            return False
        words = conf.words2 if '全军武器装备' in item['sources'] else conf.words
        for w in words:
            if w in res.text:
                return True
        return False


class MongoDB_pipline(object):
    def process_item(self, item, spider):
        redis_cli = redis.Redis(host="127.0.0.1", port=6379, db="0")
        mongo_cli = pymongo.MongoClient(host="127.0.0.1", port=27017)
        # 创建mongodb数据库名称
        dbname = mongo_cli["student"]
        # 创建mongodb数据库的表名称
        sheet_name = dbname["beijing"]
        offset = 0
        while True:
            source, data = redis_cli.blpop("yy:items")
            offset += 1
            data = json.loads(data)
            sheet_name.insert(data)
            print(offset)


class MySQL_pipline(object):
    def process_item(self, item, spider):
        redis_cli = redis.Redis(host="127.0.0.1", port=6379, db=0)
        mysql_cli = pymysql.connect(host="127.0.0.1", port=3306, user="admin", passwd="123456", db="student")
        offset = 0
        while True:
            source, data = redis_cli.blpop("yy:items")
            item = json.loads(data)
            sql = "insert into beijing (username, age, header_url) values (%s, %s, %s)"
            try:
                cursor = mysql_cli.cursor()
                cursor.execute(sql, [item['username'], item['age'], item['header_url']])
                mysql_cli.commit()
                cursor.close()
                offset += 1
                print(offset)
            except:
                pass

